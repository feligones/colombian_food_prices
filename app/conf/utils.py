import pandas as pd
import numpy as np
import pickle
import requests
from openpyxl import load_workbook
import os
from unidecode import unidecode
from nltk.tokenize import word_tokenize
import re
from zeep import Client
from zeep.helpers import serialize_object

# Function to save an object as a pickle in a specified local path
def dump_artifact(artifact, artifact_name, path):
    """
    This method saves an object as a pickle in a specified local path.

    Args:
        artifact (object): object to be saved as an artifact.
        artifact_name (string): name of the saved artifact.
        path (string): local path to save artifacts.
    """
    with open(path + artifact_name + '.pkl', 'wb') as handle:
        pickle.dump(artifact, handle, protocol=pickle.HIGHEST_PROTOCOL)

# Function to load a previously saved artifact
def load_artifact(artifact_name, path):
    """
    This method returns the queried artifact in its original form.

    Args:
        artifact_name (string): name of the saved artifact.
        path (string): local path to collect artifacts.

    Returns:
        artifact (object): object saved as the queried artifact.
    """
    with open(path + artifact_name + '.pkl', 'rb') as handle:
        artifact = pickle.load(handle)
    return artifact

# Function to load a consolidated prices dataframe from an XLSX file
def load_dataset(filename, local_path, url_path):
    """
    This method returns a consolidated prices dataframe.

    The process consists of requesting an XLSX file from the DANE webpage 
    corresponding to one or multiple years of monthly average prices.

    Args:
        filename (string): name of the XLSX file.
        local_path (string): local path to save the requested file.
        url_path (string): remote path to request the XLSX file.

    Returns:
        concat_dataframe (pd.Dataframe): pandas dataframe form of the 
                                         requested XLSX file.
    """
    # Create path to save the requested XLSX file
    file_path = local_path + filename

    # Check if the file is not already saved locally
    if filename not in os.listdir(local_path):
        # Create the complete request URL
        url = url_path + filename
        # Request the file from the URL
        resp = requests.get(url)
        assert resp.status_code == 200, resp.status_code + " - " + filename
        # Save the requested file locally
        file = open(file_path, 'wb')
        file.write(resp.content)
        file.close()

    # Instance a workbook
    workbook = load_workbook(file_path, data_only=True)
    # Get sheet names from the workbook
    sheetnames = workbook.sheetnames

    _dataframe_list = []
    # For each sheetname...
    for sheet in sheetnames[1:]:
        # Create a pandas DataFrame with values
        dataframe = pd.DataFrame(workbook[sheet].values)
        # Clean useless rows and columns
        _valid_index = pd.to_datetime(dataframe[0], errors='coerce').dropna(axis=0).index
        dataframe = dataframe.filter(_valid_index, axis=0).dropna(axis=1)
        _dataframe_list.append(dataframe)

    # Concatenate all dataframes of each sheet
    concat_dataframe = pd.concat(_dataframe_list, ignore_index=True)

    return concat_dataframe

# Function to load the dataset of the last month from a SOAP service
def load_last_month_dataset(cutoff_date):
    """
    This method returns the dataset of the last month from a SOAP service.

    Args:
        cutoff_date (datetime): cutoff date for the data.

    Returns:
        month_price_data (pd.Dataframe): pandas dataframe containing the data.
    """
    client = Client("https://appweb.dane.gov.co/sipsaWS/SrvSipsaUpraBeanService?WSDL")

    week_price_data = pd.DataFrame(serialize_object(client.service.promediosSipsaSemanaMadr()))
    week_price_data['fechaIni'] = pd.to_datetime(week_price_data['fechaIni'].astype(str).str.split(' ', expand=True)[0])

    rename_columns = {
        'fechaIni': 'date',
        'artiNombre': 'product',
        'fuenNombre': 'market',
        'promedioKg': 'mean_price'
    }
    select_columns = list(rename_columns.values())

    week_price_data = week_price_data.loc[(week_price_data['fechaIni'] >= cutoff_date)].copy()
    week_price_data = week_price_data.rename(columns=rename_columns)[select_columns]

    month_price_data = week_price_data.groupby([pd.Grouper(key='date', freq='MS'), 'product', 'market'])['mean_price'].mean().reset_index()

    return month_price_data

# Function to clean text by removing accents, special characters, and lowercasing
def clean_text(text):
    """
    This method returns the input text without accents, 
    special characters, and lowercased.

    Args:
        text (string): input text.

    Returns:
        clean_text (string): processed text.
    """
    clean_text = unidecode(text.lower())
    clean_text = re.sub(r'[^\w\s]', '', clean_text)
    clean_text = ' '.join(word_tokenize(clean_text))
    return clean_text